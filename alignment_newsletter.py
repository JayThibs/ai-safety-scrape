import os
import re
import pandas as pd
import openpyxl
import requests
import jsonlines
from utils import *
import arxiv
from bs4 import BeautifulSoup as bs
from markdownify import MarkdownConverter
from paper2json.tei2json import extract_body_as_markdown_from_tei
from tqdm import tqdm
import multiprocessing as mp
import concurrent.futures


class AlignmentNewsletter:
    def __init__(
        self,
        n_threads=1,
    ):
        self.n_threads = n_threads
        self.setup()

    def fetch_entries(self):
        print("Fetching alignment_newsletter entries")
        self.alignment_newsletter = {}
        self.df = self.df.iloc[0:5]
        # with concurrent.futures.ProcessPoolExecutor(self.n_threads) as executor:
        #     executor.map(self.fetch_individual_entries, self.df.to_dict("records"))
        for index, row in self.df.iterrows():
            self.fetch_individual_entries(index, row)

        print("Removing entries with no summaries...")
        alignment_newsletter_entries = {
            k: v
            for k, v in self.alignment_newsletter.items()
            if v["paper_summary"] != "nan"
        }
        alignment_newsletter_entry_list = []
        for entry in alignment_newsletter_entries.keys():
            alignment_newsletter_entry_list.append(alignment_newsletter_entries[entry])

        print("Creating jsonl and txt file...")
        for i, entry in enumerate(alignment_newsletter_entry_list):
            i = str(i)
            with jsonlines.open(
                "data/alignment_newsletter_separate_summaries.jsonl", "a"
            ) as writer:
                writer.write(entry)
            with open("data/alignment_newsletter_separate_summaries.txt", "a") as f:
                # Save the entry in plain text, mainly for debugging
                text = (
                    "    ".join(("\n" + entry["text"].lstrip()).splitlines(True)) + "\n"
                )
                f.write(f"[ENTRY {i}] {text}")

        # Creating new json for each individual newsletter rather than individual summary
        alignment_newsletter = {}
        for i, entry in enumerate(alignment_newsletter_entry_list):
            newsletter_number = alignment_newsletter_entry_list[i]["newsletter_number"]
            if newsletter_number not in alignment_newsletter:
                alignment_newsletter[newsletter_number] = entry
                alignment_newsletter[newsletter_number].pop("individual_summary")

        alignment_newsletter_entry_list = []
        for entry in alignment_newsletter_entries.keys():
            alignment_newsletter_entry_list.append(alignment_newsletter_entries[entry])

        for i, entry in enumerate(alignment_newsletter_entry_list):
            i = str(i)
            with jsonlines.open("data/alignment_newsletter.jsonl", "a") as writer:
                writer.write(entry)
            with open("data/alignment_newsletter.txt", "a") as f:
                # Save the entry in plain text, mainly for debugging
                text = (
                    "    ".join(("\n" + entry["text"].lstrip()).splitlines(True)) + "\n"
                )
                f.write(f"[ENTRY {i}] {text}")

    def fetch_individual_entries(self, index, row):
        print(f"Processing entry {index}/{len(self.df)}")
        try:
            # index starts at 2 because the row 1 is the header
            paper_url = str(self.ws.cell(row=index + 2, column=3).hyperlink.target)
            newsletter_url = str(self.ws.cell(row=index + 2, column=8).hyperlink.target)
            print(newsletter_url)
        except:
            paper_url = ""
            newsletter_url = ""
            pass
        if "gradientscience" in paper_url:
            r = requests.get(paper_url)
            arxiv_id = re.findall(r"(?:arXiv:|abs/)(\d{4}\.\d{4,5})", r.text)[0]
        elif "arxiv.org" in paper_url:
            try:
                arxiv_id = re.findall(r"(?:arXiv:|abs/)(\d{4}\.\d{4,5})", paper_url)[0]
            except:
                arxiv_id = None
                pass
        else:
            arxiv_id = None

        abs = ""
        markdown_text = ""
        newsletter_text = ""
        if row["Venue"] == "arXiv":
            try:
                paper = arxiv.Search(id_list=[arxiv_id], max_results=1)
                paper = next(paper.results())
                abstract = paper.summary.replace("\n", " ")
                abs = "Paper abstract: " + abstract + "\n"
            except:
                pass
        elif row["Venue"] == "Distill":
            try:
                r = requests.get(paper_url)
                soup = bs(r.text)
                markdown_text = extract_body_as_markdown_from_tei(soup)
            except:
                pass
        # Extracting the markdown text from the newsletter
        try:
            print("Fetching newsletter text")
            r = requests.get(newsletter_url)
            soup = bs(r.text)
            newsletter_text = extract_body_as_markdown_from_tei(soup)
            newsletter_text = (
                "Highlights\n\n"
                + newsletter_text.split("Highlights\n")[1].split(" |\n")[0]
            )
        except:
            pass
        print(newsletter_text)
        summary = (
            "Title: "
            + str(row["Title"])
            + "\n"
            + "Authors: "
            + str(row["Authors"])
            + "\n"
            + str(abs)
            + "Summary: "
            + str(row["Summary"])
            + "\n"
            + "My opinion: "
            + str(row["My opinion"])
        )
        self.alignment_newsletter[str(index)] = {
            "source": str(row["Venue"]),
            "newsletter_category": str(row["Category"]),
            "highlight": True if row["Highlight?"] == "Highlight" else False,
            "newsletter_number": str(row["Email"]),
            "newsletter_url": str(newsletter_url),
            "summarizer": str(row["Summarizer"]),
            "paper_summary": str(row["Summary"]),
            "opinion": str(row["My opinion"]),
            "prerequisites": str(row["Prerequisites"]),
            "read_more": str(row["Read more"]),
            "paper_version": str(paper.get_short_id()) if abs != "" else None,
            "arxiv_id": str(arxiv_id),
            "post_title": str(paper.title) if abs != "" else str(row["Title"]),
            "authors": [str(x) for x in paper.authors]
            if abs != ""
            else str(row["Authors"]),
            "date_published": str(paper.published) if abs != "" else row["Year"],
            "data_last_modified": str(paper.updated) if abs != "" else "",
            "url": str(paper.entry_id) if abs != "" else paper_url,
            "abstract": str(abstract) if abs != "" else "",
            "author_comment": str(paper.comment) if abs != "" else "",
            "journal_ref": str(paper.journal_ref) if abs != "" else "",
            "doi": str(paper.doi) if abs != "" else "",
            "primary_category": str(paper.primary_category) if abs != "" else "",
            "categories": str(paper.categories) if abs != "" else "",
            "individual_summary": str(summary),
            "paper_text": str(markdown_text),
            "text": newsletter_text,
            "bibliography_bbl": "",
            "bibliography_bib": "",
        }

    def setup(self):
        sh("mkdir -p data/processed/alignment_newsletter")
        if os.path.exists("data/alignment_newsletter.jsonl"):
            print("Deleting old alignment_newsletter.jsonl")
            os.remove("data/alignment_newsletter.jsonl")
            os.remove("data/alignment_newsletter_separate_summaries.jsonl")
        if os.path.exists("data/aligned_newsletter.txt"):
            os.remove("data/alignment_newsletter.txt")
            os.remove("data/alignment_newsletter_separate_summaries.txt")
        # put the alignment_newsletter.xlsx file in the raw/alignment_newsletter folder
        # download new excel file here: https://docs.google.com/spreadsheets/d/1PwWbWZ6FPqAgZWOoOcXM8N_tUCuxpEyMbN1NYYC02aM/edit#gid=0
        self.df = pd.read_excel(
            "data/raw/alignment_newsletter/alignment_newsletter.xlsx"
        )
        self.df["index"] = self.df.index
        wb = openpyxl.load_workbook(
            "data/raw/alignment_newsletter/alignment_newsletter.xlsx"
        )
        self.ws = wb["Sheet1"]


if __name__ == "__main__":

    AlignmentNewsletter(n_threads=mp.cpu_count()).fetch_entries()
